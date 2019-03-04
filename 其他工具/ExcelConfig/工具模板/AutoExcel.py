#coding:utf-8
import time
#import xlrd
from openpyxl import load_workbook
from shutil import copyfile
import csv
import datetime
import sys
import os
reload(sys)

#UTF-8 无BOM格式
sys.setdefaultencoding('utf_8')
#UTF-8 带BOM格式
#sys.setdefaultencoding('utf_8_sig')

DataPathCSV = "Client\Assets\AssetResources\ExternalResource"
DataPathXlsm = "ExcelConfig\工具模板"

# 读取config.txt中需要打包的excel文件
def Read_config():

	alltime_start = datetime.datetime.now()
	errCSV=[]
	print "开始生成配置文件：".encode('gbk')
	
	#print SetPath(DataPathXlsm,'CSV转成XLSM.xlsm').encode('gbk')
	copyfile(SetPath(DataPathXlsm,'CSV转成XLSM.xlsm').encode('gbk'),'CSV转成XLSM.xlsm'.encode('gbk'))
	csv_xlsm = FindXlsm(GetFilesName(SetPath(DataPathCSV,'CSV'),"csv"))
	for csv_xlsm_file in csv_xlsm:
		try:
			#print "csvfile",csvfile
			#print SetPath(DataPathCSV,'CSV')+"\\"+csv_xlsm_file[0]
			config = csv.reader(open(SetPath(DataPathCSV,'CSV')+"\\"+csv_xlsm_file[0],'r'))
			# 打开文件
			starttime = datetime.datetime.now()
			#print csv_xlsm_file[1]
			try:
				wb = load_workbook(csv_xlsm_file[1],keep_vba=True)
			except:
				wb = load_workbook('CSV转成XLSM.xlsm'.encode('gbk'),keep_vba=True)
		
			#wb.create_sheet('Sheet2', index=1)  # 被安排到第二个工作表，index=0就是第一个位置
			sheetname = wb.get_sheet_names()
			ws = wb.get_sheet_by_name(sheetname[0])
			#ws2 = wb.get_sheet_by_name(sheetname[1])
			#wb.remove(ws2)
		
			row = ws.max_row
			col = ws.max_column
			#print row,' ',col
			for i in range(1,row+1):
				for j in range(1,col+1):
					ws.cell(i,j).value = None

			i = 0
			for line in config:
				#print line
				i += 1
				for j in range(0,len(line)):
					ws.cell(i,j+1).value = line[j]

			#print row,' ',col
			#for line in config:
			#	ws.append(line)
			wb.save(csv_xlsm_file[1])
			endtime = datetime.datetime.now()-starttime
			strtimes = csv_xlsm_file[1] + Autospace(60,len(csv_xlsm_file[1])) + str(endtime)
			print strtimes
		except:
			errCSV.append(csv_xlsm_file[0])
	alltime_end = datetime.datetime.now()-alltime_start
	print "总耗时：".encode('gbk') + str(alltime_end) + '\n'
	if len(errCSV)>0:
		Writefile('errCSV',errCSV,os.getcwd(),'.txt')
		print '本次有未处理成功的文件，请查看本目录下的"errCSV"文件'.encode('gbk')
		time.sleep(10)

		
#存储文本内容到文本文件（文件名，文件内容，保存路径，后缀名）
def Writefile(test_name,test,path,endname):

	if findstr(test_name,"_"):
		test_name=str(test_name[test_name.find("_")+1:]).encode('gbk')
	else:
		test_name=str(test_name)
		
	#当不存在对应的文件夹时，自动创建对应的文件夹
	if not os.path.exists(path):
		os.makedirs(path)

	test_name = path + "\\" + test_name + endname
	test = str(test)
	file = open(test_name,"w")
	file.write(str(test))
	file.close()


def Autospace(a,b):
	space=""
	for i in range(a-b):
		space = space + " "
	return space
	

def findstr(a,b):
	a=a.lower()
	b=b.lower()
	if a.find(b)>=0:
		return True
	else:
		return False

		
#自动设置csv文件的获取路径
def SetPath(DataPath,dirname):
	path=os.getcwd()
	pathlist=path.rsplit("\\")
	lenth = path.find("ExcelConfig")
	path = path[0:lenth]
	if dirname == 'CSV':
		path = path + DataPath + "\\" + pathlist[-1] + "\\"+dirname
	else:
		path = path + DataPath + "\\"+dirname
	#print path.encode('gbk')
	return path
	

#根据路径进行存储数据文件
def GetFilePath(filename):
	path=""
	config=open(filename,"r")
	for line in open(filename,"r"):
		line = config.readline()
		#print line
		if findstr(line,"csv"):
			path=line.split('=')[1].rstrip('\n').rstrip(' ')
	return path


#获取"firl_dir"路径下所有后缀名为"file_endname"的文件
def GetFilesName(firl_dir,file_endname):
	files = os.listdir(firl_dir)
	files_name=[]
	lenth=len(file_endname)-file_endname.find('.')
	for line in files:
		if len(file_endname)>0 and line[-lenth:] =="."+file_endname:
			files_name.append(line)
	return files_name

	
#找到本CSV文件对应的xlsm文件（有中文备注的也会自动找到）
def FindXlsm(files_name):
	xlsm_dir = os.getcwd()
	files_xlsm = GetFilesName(xlsm_dir,'xlsm')
	xlsm_lists=[]
	csv_xlsm=[]
	for line in files_xlsm:
		if line.find('_')>0:
			xlsmlist = line.rsplit('_')
			xlsm_lists.append(xlsmlist)
		else:
			xlsmlist = [[],line]
			xlsm_lists.append(xlsmlist)
	#print xlsm_lists
	
	for csvname in files_name:
		x=0
		csvname = csvname.replace('.csv','.xlsm')
		for xlsmname in xlsm_lists:
			#print csvname,'    ',xlsmname[1]
			if csvname == xlsmname[1]:
				if xlsmname[0]==[]:
					csv_xlsm.append([csvname.replace('.xlsm','.csv'),xlsmname[1]])
				else:
					csv_xlsm.append([csvname.replace('.xlsm','.csv'),xlsmname[0]+'_'+xlsmname[1]])
				x+=1
				continue
		if x == 0:
			csv_xlsm.append([csvname.replace('.xlsm','.csv'),csvname])
	#print csv_xlsm
	return csv_xlsm


#if __name__ == '__main__':
if __name__=='__main__':

	#FindXlsm(GetFilesName(SetPath(DataPathCSV,'CSV'),"csv"))
	Read_config()
	os.remove('CSV转成XLSM.xlsm'.encode('gbk'))
	#print SetPath(DataPathCSV,'CSV')
	#SetPath(DataPathXlsm,'CSV转成XLSM.xlsm')